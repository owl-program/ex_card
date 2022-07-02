from cProfile import label
import openpyxl
import tkinter
import tkinter.ttk as ttk

root = tkinter.Tk()
root.title('Excell_Card')
root.geometry('600x250')

wb = openpyxl.load_workbook('card.xlsx')
ws = wb['Sheet1']

#タイトルパーツ
title_v = tkinter.Entry(root, width=30)
label_t = tkinter.Label(root, text="タイトル")
label_t.pack()
title_v.pack()

#プルダウン式の言語パーツ
list = ('Java', 'VBA', 'Python', 'javaScript', 'PHP', 'C', 'Django', 'Laravel', 'SQL')
lang = ttk.Combobox(root, values=list)
label_l = tkinter.Label(root, text="言語区分")
label_l.pack()
lang.pack()

#概要パーツ
summary = tkinter.Entry(root, width=100)
label_s = tkinter.Label(root, text="概要欄")
label_s.pack()
summary.pack()

#活用例パーツ
use_v = tkinter.Entry(root, width=100)
label_u = tkinter.Label(root, text="活用例")
label_u.pack()
use_v.pack()

#キーワードパーツ
key_word = tkinter.Entry(root, width=50)
label_k = tkinter.Label(root, text="キーワード")
label_k.pack()
key_word.pack()

def word_card():
    maxRow = ws.max_row + 1
    Value_v = title_v.get()
    Value_l = lang.get()
    Value_s = summary.get()
    Value_u = use_v.get()
    Value_k = key_word.get()
    ws.cell(row=maxRow, column=1).value = Value_v
    ws.cell(row=maxRow, column=2).value = Value_l
    ws.cell(row=maxRow, column=3).value = Value_s
    ws.cell(row=maxRow, column=4).value = Value_u
    ws.cell(row=maxRow, column=5).value = Value_k
    wb.save('card.xlsx')

def clearTextInput():
    title_v.delete(0, tkinter.END)
    lang.delete(0, tkinter.END)
    summary.delete(0, tkinter.END)
    use_v.delete(0, tkinter.END)
    key_word.delete(0, tkinter.END)

clear_b=tkinter.Button(root, height=1, width=10, text="クリア", command=clearTextInput)
clear_b.pack()

registration = tkinter.Button(root, text = '登録', command = word_card)
registration.pack()

print(ws.max_column)
print(ws.max_row)
root.mainloop()